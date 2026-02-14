import os


import openpyxl
# kod 2023 için gecerli subat 28 29 cekmesıne göre değişir
# sehir esitlemesi falan hepsi düzeltilmeli yıl degisince !!!
#-----------------------------------------------------------------------------





#-----------------------------------------------------------------------------



yil ='2026'

satirlar = []                 # 365 gun ıcın 365 satır alır 0 - 364 indexinde
kible_saati="00 00"           # kalip değer kıble saati için
start_row       = 2           # excall deki baslangıc satır numarası
start_column    = 2           # excall deki baslangıc sutun numarası
sehir_bosluk_bas="                              "
sehir_bosluk_son="           \n"







excell_files = [f for f in os.listdir('.') if f.endswith('.xlsx')]
if len(excell_files) != 1:
    raise ValueError('EXCELL DOSYASI DIRECTORYDE YOK!!')

filename = excell_files[0]

# excell dosyasını ac
wb = openpyxl.load_workbook(filename)
sheet = wb.active
# kalip txt dosyasını ac
with open('vakit_kalip.txt', 'r') as file:
    data = file.readlines()



sehir_adi = filename[:-12]

sehir = sehir_bosluk_bas + sehir_adi + sehir_bosluk_son

data[4] = sehir
data[5] ="                             "+yil+" - OCAK\n"
data[40] = sehir
data[41] ="                             "+yil+" - SUBAT\n"
data[73] = sehir
data[74] ="                             "+yil+" - MART\n"
data[109] = sehir
data[110] ="                             "+yil+" - NISAN\n"
data[144] = sehir
data[145] ="                             "+yil+" - MAYIS\n"
data[180] = sehir
data[181] ="                             "+yil+" - HAZIRAN\n"
data[215] = sehir
data[216] ="                             "+yil+" - TEMMUZ\n"
data[251] = sehir
data[252] ="                             "+yil+" - AGUSTOS\n"
data[287] = sehir
data[288] ="                             "+yil+" - EYLUL\n"
data[322] = sehir
data[323] ="                             "+yil+" - EKIM\n"
data[358] = sehir
data[359] ="                             "+yil+" - KASIM\n"
data[393] = sehir
data[394] ="                             "+yil+" - ARALIK\n"


for x in range(365):

    imsak = str(sheet.cell(row=start_row,column=start_column).value)
    # imsak =imsak[:-3]
    imsak = imsak.replace(":", " ")
    gunes = str(sheet.cell(row=start_row,column=start_column+1).value)
    # gunes =gunes[:-3]
    gunes = gunes.replace(":", " ")
    ogle = str(sheet.cell(row=start_row,column=start_column+2).value)
    # ogle =ogle[:-3]
    ogle = ogle.replace(":", " ")
    ikindi = str(sheet.cell(row=start_row,column=start_column+3).value)
    # ikindi =ikindi[:-3]
    ikindi = ikindi.replace(":", " ")
    aksam = str(sheet.cell(row=start_row,column=start_column+4).value)
    # aksam =aksam[:-3]
    aksam = aksam.replace(":", " ")
    yatsi = str(sheet.cell(row=start_row,column=start_column+5).value)
    # yatsi =yatsi[:-3]
    yatsi = yatsi.replace(":", " ")

    satirlar.append("   "+imsak+"  "+gunes+"  "+ogle+"  "+ikindi+"  "+aksam+"  "+yatsi+"  "+kible_saati)

    start_row = start_row + 1



def ay_duzenle(pos1,pos2):

    gun_ct = 1 # gun counter
    pos_ct = 0 # pos tutuyor sonra ay bitis charı ekliyecek
    for x in range (pos1,pos2):
        satirlar[x]="           " + "{:02d}".format(gun_ct) + satirlar[x]
        gun_ct = gun_ct + 1
        pos_ct = x

    satirlar[pos_ct]=satirlar[pos_ct] + ''


ay_duzenle(0,31) # ocak ayı
ay_duzenle(31,59) # subat ayı
ay_duzenle(59,90) # mart ayı
ay_duzenle(90,120) # nisan ayı
ay_duzenle(120,151) # mayıs ayı
ay_duzenle(151,181) # haziran ayı
ay_duzenle(181,212) # temmuz ayı
ay_duzenle(212,243) # agustos ayı
ay_duzenle(243,273) # eylül ayı
ay_duzenle(273,304) # ekim ayı
ay_duzenle(304,334) # kasım ayı
ay_duzenle(334,365) # aralık ayı

for x in range(0,365):
    satirlar[x] = satirlar[x] + '\n'




def ay_yaz(_index,_pos1,_pos2):
    for x in range (_pos1,_pos2):
        data[_index] = satirlar[x]
        _index += 1



ay_yaz(8,0,31) # ocak ayi 
ay_yaz(44,31,59) # subat ayi 
ay_yaz(77,59,90) # mart ayi 
ay_yaz(113,90,120) # nisan ayi 
ay_yaz(148,120,151) # mayıs ayi 
ay_yaz(184,151,181) # haziran ayi 
ay_yaz(219,181,212) # temmuz ayi 
ay_yaz(255,212,243) # agustos ayi 
ay_yaz(291,243,273) # eylul ayi 
ay_yaz(326,273,304) # ekim ayi 
ay_yaz(362,304,334) # kasım ayi 
ay_yaz(397,334,365) # aralık ayi 


with open(sehir_adi+'.txt', 'w') as file:
    file.writelines( data )


# start_index = 8  # ocak ayi txt dosyasında 8. indexte
# for x in range (0,31):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 44  # subat ayi ilk gun index
# for x in range (31,59):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 77  # mart ayi ilk gun index
# for x in range (59,90):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 113  # nisan ayi ilk gun index
# for x in range (90,120):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 148  # mayıs ayi ilk gun index
# for x in range (120,151):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 184  # haziran ayi ilk gun index
# for x in range (151,181):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 219  # temmuz ayi ilk gun index
# for x in range (181,212):
#     data[start_index] = satirlar[x]
#     start_index += 1 
# start_index = 255  # agustos ayi ilk gun index
# for x in range (212,243):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 291  # eylul ayi ilk gun index
# for x in range (243,273):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 326  # ekim ayi ilk gun index
# for x in range (273,304):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 362  # kasim ayi ilk gun index
# for x in range (304,334):
#     data[start_index] = satirlar[x]
#     start_index += 1
# start_index = 397  # aralik ayi ilk gun index
# for x in range (334,365):
#     data[start_index] = satirlar[x]
#     start_index += 1


# with open('vakit_cikti.txt', 'w') as file:
#     file.writelines( data )








# data[8] =satirlar[0]
# # print(data[8])
# with open('vakit_kalip.txt', 'w') as file:
#     file.writelines( data )


# # with is like your try .. finally block in this case
# with open('vakit_kalip.txt', 'r') as file:
#     # read a list of lines into data
#     data = file.readlines()

# data[8] = "           01   09 49  08 21  13 12  15 32  17 53  19 19  11 12\n"
# # print(data[8])
# with open('vakit_kalip.txt', 'w') as file:
#     file.writelines( data )





# # with is like your try .. finally block in this case
# with open('vakit_kalip.txt', 'r') as file:
#     # read a list of lines into data
#     data = file.readlines()
# # sabah 16,17-19,20
# # print(data[8])
# position = 17
# data[8] = data[8][:position] + '7' + data[8][position+1:]
# # print(data[8])
# with open('vakit_kalip.txt', 'w') as file:
#     file.writelines( data )